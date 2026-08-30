import tkinter as tk
import ttkbootstrap as tb
from ttkbootstrap.constants import *
from ttkbootstrap.dialogs import Messagebox
from tkinter import filedialog, Toplevel, Listbox
import sqlite3
import csv
import json
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
import os
# ===================== 自动补全组件 =====================
class AutocompleteEntry(tb.Frame):
    def __init__(self, parent, width=26, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)
        self.parent = parent
        self.width = width
        self.var = tk.StringVar()
        self.entry = tb.Entry(self, textvariable=self.var, width=width)
        self.entry.pack(fill="x")
        self.list_win = None
        self.listbox = None
        self.item_list = []
        self.selected_value = None
        self.var.trace_add("write", self.on_text_change)
        self.entry.bind("<Down>", self.focus_list)
        self.entry.bind("<Escape>", self.hide_list)
    def set_items(self, item_list):
        self.item_list = item_list
    def get(self):
        return self.selected_value
    def set(self, text):
        self.var.set(text)
        self.selected_value = text
    def on_text_change(self, *args):
        text = self.var.get().lower()
        if not text:
            self.hide_list()
            return
        matches = [i for i in self.item_list if text in i.lower()]
        if matches:
            self.show_list(matches[:12])
        else:
            self.hide_list()
    def show_list(self, matches):
        if self.list_win is None or not self.list_win.winfo_exists():
            self.list_win = Toplevel(self)
            self.list_win.overrideredirect(True)
            self.list_win.attributes("-topmost", True)
            self.listbox = Listbox(self.list_win, width=self.width+8, height=8)
            self.listbox.pack()
            self.listbox.bind("<<ListboxSelect>>", self.on_select_item)
            self.listbox.bind("<Escape>", self.hide_list)
        x = self.entry.winfo_rootx()
        y = self.entry.winfo_rooty() + self.entry.winfo_height()
        self.list_win.geometry(f"+{x}+{y}")
        self.listbox.delete(0, tk.END)
        for m in matches:
            self.listbox.insert(tk.END, m)
        self.list_win.deiconify()
    def hide_list(self, event=None):
        if self.list_win and self.list_win.winfo_exists():
            self.list_win.destroy()
        self.list_win = None
        self.listbox = None
    def focus_list(self, event):
        if self.listbox:
            self.listbox.focus_set()
            self.listbox.selection_set(0)
        return "break"
    def on_select_item(self, event):
        idx = self.listbox.curselection()
        if idx:
            val = self.listbox.get(idx[0])
            self.var.set(val)
            self.selected_value = val
            self.hide_list()
# ===================== 加载外部配置文件 =====================
CONFIG_PATH = "config.json"
def load_config():
    if not os.path.exists(CONFIG_PATH):
        raise FileNotFoundError("请把config.json和程序放在同一文件夹！")
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        return json.load(f)
CFG = load_config()
STAGE_LIST = CFG["customer_stage_options"]
FOLLOW_TYPE_LIST = CFG["follow_type_options"]
SEARCH_FIELDS = CFG["search_db_fields"]
# ===================== 数据库工具函数 =====================
def init_db():
    conn = sqlite3.connect("crm.db")
    cur = conn.cursor()
    cur.execute("""CREATE TABLE IF NOT EXISTS customer (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT,
        legal_person TEXT,
        phone TEXT,
        address TEXT,
        stage TEXT,
        business_scope TEXT,
        registered_cap TEXT,
        remark TEXT,
        tags TEXT
    )""")
    cur.execute("""CREATE TABLE IF NOT EXISTS follow_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        cust_id INTEGER,
        follow_type TEXT,
        follow_date TEXT,
        next_remind TEXT,
        remark TEXT,
        FOREIGN KEY(cust_id) REFERENCES customer(id)
    )""")
    try:
        cur.execute("ALTER TABLE customer ADD COLUMN tags TEXT")
    except:
        pass
    try:
        cur.execute("ALTER TABLE customer ADD COLUMN business_scope TEXT")
    except:
        pass
    try:
        cur.execute("ALTER TABLE customer ADD COLUMN registered_cap TEXT")
    except:
        pass
    try:
        cur.execute("ALTER TABLE customer ADD COLUMN remark TEXT")
    except:
        pass
    conn.commit()
    conn.close()
def db_execute(sql, params=()):
    conn = sqlite3.connect("crm.db")
    cur = conn.cursor()
    try:
        cur.execute(sql, params)
        conn.commit()
    except Exception as e:
        conn.rollback()
    finally:
        conn.close()
def db_query(sql, params=()):
    conn = sqlite3.connect("crm.db")
    cur = conn.cursor()
    try:
        cur.execute(sql, params)
        res = cur.fetchall()
        return res
    except Exception as e:
        return []
    finally:
        conn.close()
def get_table_columns(table_name):
    conn = sqlite3.connect("crm.db")
    cur = conn.cursor()
    cur.execute(f"PRAGMA table_info({table_name})")
    res = cur.fetchall()
    conn.close()
    return [row[1] for row in res]
def add_col_if_not_exist(table_name, col_name):
    cols = get_table_columns(table_name)
    if col_name in cols:
        return
    conn = sqlite3.connect("crm.db")
    cur = conn.cursor()
    cur.execute(f"ALTER TABLE {table_name} ADD COLUMN `{col_name}` TEXT")
    conn.commit()
    conn.close()
# ===================== 工具：生成带滚动的Toplevel弹窗 =====================
def create_scrolled_toplevel(root, title, w=700, h=550):
    win = Toplevel(root)
    win.title(title)
    win.geometry(f"{w}x{h}")
    win.resizable(True, True)
    win.grab_set()
    win.transient(root)
    canvas = tk.Canvas(win)
    v_scroll = tb.Scrollbar(win, orient=VERTICAL, command=canvas.yview)
    h_scroll = tb.Scrollbar(win, orient=HORIZONTAL, command=canvas.xview)
    canvas.configure(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)
    v_scroll.pack(side=RIGHT, fill=Y)
    h_scroll.pack(side=BOTTOM, fill=X)
    canvas.pack(side=LEFT, fill=BOTH, expand=True)
    inner_frame = tb.Frame(canvas)
    canvas.create_window((0,0), window=inner_frame, anchor="nw")
    def _on_frame(event):
        canvas.configure(scrollregion=canvas.bbox("all"))
    inner_frame.bind("<Configure>", _on_frame)
    def _mousewheel(event):
        canvas.yview_scroll(int(-1*(event.delta/120)), "units")
    canvas.bind_all("<MouseWheel>", _mousewheel)
    return win, canvas, inner_frame
# ====================== GUI主程序 ======================
class SalesCRM:
    def __init__(self, root):
        self.root = root
        self.root.title("KCRM by KZG Tech v2")
        self.root.geometry("1160x760")
        self.cust_sort_col = None
        self.cust_sort_desc = False
        self.follow_sort_col = None
        self.follow_sort_desc = False
        self.notebook = tb.Notebook(root)
        self.frame_dash = tb.Frame(self.notebook)
        self.frame_cust = tb.Frame(self.notebook)
        self.frame_follow = tb.Frame(self.notebook)
        self.notebook.add(self.frame_dash, text="🏠 首页看板")
        self.notebook.add(self.frame_cust, text="📋 客户台账")
        self.notebook.add(self.frame_follow, text="📝 跟进日志")
        self.notebook.pack(expand=True, fill="both", padx=10, pady=10)
        self.notebook.bind("<<NotebookTabChanged>>", self.on_tab_change)
        self.build_dash_tab()
        self.build_cust_tab()
        self.build_follow_tab()
        self.refresh_all()
        self.check_remind()
    def on_tab_change(self, event):
        idx = self.notebook.index(self.notebook.select())
        if idx == 0:
            self.refresh_dashboard()
        elif idx == 1:
            self.refresh_cust_table()
        elif idx == 2:
            self.reload_customer_combobox()
    def refresh_all(self):
        self.refresh_dashboard()
        self.refresh_cust_table()
        self.refresh_follow_table()
        self.reload_customer_combobox()
    def sort_treeview(self, tree, sort_col, is_desc, refresh_func):
        for col in tree["columns"]:
            tree.heading(col, text=tree.heading(col)["text"].replace(" ▲", "").replace(" ▼", ""))
        sort_mark = " ▼" if is_desc else " ▲"
        tree.heading(sort_col, text=tree.heading(sort_col)["text"] + sort_mark)
        refresh_func(sort_col, is_desc)
    def build_dash_tab(self):
        title = tb.Label(self.frame_dash, text="数据概览", font=("微软雅黑", 18, "bold"))
        title.pack(pady=(20, 10))
        card_frame = tb.Frame(self.frame_dash)
        card_frame.pack(pady=10)
        self.card_total = self._make_card(card_frame, "客户总数", "0", PRIMARY, 0)
        self.card_today = self._make_card(card_frame, "今日待回访", "0", WARNING, 1)
        self.card_tomorrow = self._make_card(card_frame, "明日待回访", "0", INFO, 2)
        self.card_done = self._make_card(card_frame, "今日已回访", "0", SUCCESS, 3)
        tb.Label(self.frame_dash, text="今日待回访明细", font=("微软雅黑", 12, "bold")).pack(pady=(20, 5))
        self.dash_tree = tb.Treeview(self.frame_dash, columns=("cust_name", "follow_type", "next_remind", "remark"), show="headings", height=8)
        self.dash_tree.heading("cust_name", text="客户名称")
        self.dash_tree.heading("follow_type", text="跟进方式")
        self.dash_tree.heading("next_remind", text="回访日期")
        self.dash_tree.heading("remark", text="备注")
        self.dash_tree.column("cust_name", width=200)
        self.dash_tree.column("follow_type", width=100)
        self.dash_tree.column("next_remind", width=120)
        self.dash_tree.column("remark", width=440)
        self.dash_tree.pack(padx=20, pady=5, fill="both", expand=True)
    def _make_card(self, parent, label, value, bootstyle, col):
        frame = tb.Frame(parent, padding=20)
        frame.grid(row=0, column=col, padx=15)
        lbl_val = tb.Label(frame, text=value, font=("微软雅黑", 28, "bold"), bootstyle=bootstyle)
        lbl_val.pack()
        tb.Label(frame, text=label, font=("微软雅黑", 11)).pack(pady=(5, 0))
        return lbl_val
    def refresh_dashboard(self):
        today = datetime.now().strftime("%Y-%m-%d")
        tomorrow = (datetime.now() + timedelta(days=1)).strftime("%Y-%m-%d")
        total = db_query("SELECT COUNT(*) FROM customer")[0][0]
        self.card_total.config(text=str(total))
        today_remind = db_query("SELECT COUNT(DISTINCT cust_id) FROM follow_log WHERE next_remind = ?", (today,))[0][0]
        self.card_today.config(text=str(today_remind))
        tom_remind = db_query("SELECT COUNT(DISTINCT cust_id) FROM follow_log WHERE next_remind = ?", (tomorrow,))[0][0]
        self.card_tomorrow.config(text=str(tom_remind))
        today_follow = db_query("SELECT COUNT(*) FROM follow_log WHERE follow_date = ?", (today,))[0][0]
        self.card_done.config(text=str(today_follow))
        for i in self.dash_tree.get_children():
            self.dash_tree.delete(i)
        rows = db_query("""
            SELECT c.name, f.follow_type, f.next_remind, f.remark
            FROM follow_log f LEFT JOIN customer c ON f.cust_id = c.id WHERE f.next_remind = ?
        """, (today,))
        for row in rows:
            self.dash_tree.insert("", tk.END, values=row)
    def build_cust_tab(self):
        search_bar = tb.Frame(self.frame_cust)
        search_bar.pack(fill="x", padx=8, pady=4)
        tb.Label(search_bar, text="多条件搜索（名称/法人/电话/跟进备注）：").pack(side="left")
        self.search_var = tk.StringVar()
        tb.Entry(search_bar, textvariable=self.search_var, width=45).pack(side="left", padx=6)
        tb.Button(search_bar, text="搜索", command=self.do_search, bootstyle=PRIMARY).pack(side="left", padx=3)
        tb.Button(search_bar, text="重置", command=self.refresh_cust_table, bootstyle=SECONDARY).pack(side="left", padx=3)
        tb.Button(search_bar, text="高级筛选", command=self.open_advance_filter, bootstyle=INFO).pack(side="left", padx=10)
        f_input = tb.Frame(self.frame_cust)
        f_input.pack(fill="x", padx=8, pady=4)
        tb.Label(f_input, text="企业名称").grid(row=0, column=0, padx=3)
        self.var_cust_name = tk.StringVar()
        tb.Entry(f_input, textvariable=self.var_cust_name, width=18).grid(row=0, column=1, padx=3)
        tb.Label(f_input, text="法人").grid(row=0, column=2, padx=3)
        self.var_legal = tk.StringVar()
        tb.Entry(f_input, textvariable=self.var_legal, width=14).grid(row=0, column=3, padx=3)
        tb.Label(f_input, text="电话").grid(row=0, column=4, padx=3)
        self.var_phone = tk.StringVar()
        tb.Entry(f_input, textvariable=self.var_phone, width=14).grid(row=0, column=5, padx=3)
        f_input2 = tb.Frame(self.frame_cust)
        f_input2.pack(fill="x", padx=8, pady=4)
        tb.Label(f_input2, text="地址").grid(row=0, column=0, padx=3)
        self.var_addr = tk.StringVar()
        tb.Entry(f_input2, textvariable=self.var_addr, width=32).grid(row=0, column=1, padx=3)
        tb.Label(f_input2, text="客户阶段").grid(row=0, column=2, padx=3)
        self.var_stage = tk.StringVar()
        tb.Combobox(f_input2, textvariable=self.var_stage, values=STAGE_LIST, width=14, state="readonly").grid(row=0, column=3, padx=3)
        tb.Label(f_input2, text="标签").grid(row=0, column=4, padx=3)
        self.var_tag = tk.StringVar()
        tb.Entry(f_input2, textvariable=self.var_tag, width=16).grid(row=0, column=5, padx=3)
        f_btn = tb.Frame(self.frame_cust)
        f_btn.pack(fill="x", padx=8, pady=8)
        tb.Button(f_btn, text="新增客户", command=self.add_cust, bootstyle=SUCCESS).pack(side="left", padx=4)
        tb.Button(f_btn, text="导入XLSX", command=self.import_xlsx, bootstyle=INFO).pack(side="left", padx=4)
        tb.Button(f_btn, text="导出全部CSV", command=self.export_all_csv, bootstyle=PRIMARY).pack(side="left", padx=4)
        tb.Button(f_btn, text="导出全部XLSX", command=self.export_all_xlsx, bootstyle=PRIMARY).pack(side="left", padx=4)
        tb.Button(f_btn, text="按日期导出CSV", command=self.export_by_date_csv, bootstyle=SECONDARY).pack(side="left", padx=4)
        tb.Button(f_btn, text="按日期导出XLSX", command=self.export_by_date_xlsx, bootstyle=SECONDARY).pack(side="left", padx=4)
        tb.Button(f_btn, text="按下次回访导出CSV", command=self.export_by_next_remind_csv, bootstyle=SECONDARY).pack(side="left", padx=4)
        tb.Button(f_btn, text="按下次回访导出XLSX", command=self.export_by_next_remind_xlsx, bootstyle=SECONDARY).pack(side="left", padx=4)
        tb.Button(f_btn, text="删除选中客户", command=self.delete_customer, bootstyle=DANGER).pack(side="left", padx=4)
        table_cols_def = [x for x in CFG["customer_table_columns"] if x["visible"]]
        col_ids = [item["field"] for item in table_cols_def]
        
        # 客户台账表格容器（增加滚动条）
        cust_table_container = tb.Frame(self.frame_cust)
        cust_table_container.pack(expand=True, fill="both", padx=8, pady=6)
        
        self.cust_tree = tb.Treeview(cust_table_container, columns=col_ids, show="headings")
        # 纵向滚动条
        cust_v_scroll = tb.Scrollbar(cust_table_container, orient=VERTICAL, command=self.cust_tree.yview)
        # 横向滚动条
        cust_h_scroll = tb.Scrollbar(cust_table_container, orient=HORIZONTAL, command=self.cust_tree.xview)
        self.cust_tree.configure(yscrollcommand=cust_v_scroll.set, xscrollcommand=cust_h_scroll.set)
        
        # 滚动条布局
        cust_v_scroll.pack(side=RIGHT, fill=Y)
        cust_h_scroll.pack(side=BOTTOM, fill=X)
        self.cust_tree.pack(fill="both", expand=True)
        
        # 配置列信息
        for cfg_col in table_cols_def:
            col_id = cfg_col["field"]
            self.cust_tree.heading(col_id, text=cfg_col["label"])
            self.cust_tree.column(col_id, width=cfg_col["width"])
            self.cust_tree.heading(col_id, command=lambda c=col_id: self.on_cust_sort(c))
        self.cust_tree.bind("<Double-1>", self.open_customer_detail)
    def on_cust_sort(self, col):
        if self.cust_sort_col == col:
            self.cust_sort_desc = not self.cust_sort_desc
        else:
            self.cust_sort_col = col
            self.cust_sort_desc = False
        self.sort_treeview(self.cust_tree, col, self.cust_sort_desc, self.refresh_cust_table)
    def on_follow_sort(self, col):
        if self.follow_sort_col == col:
            self.follow_sort_desc = not self.follow_sort_desc
        else:
            self.follow_sort_col = col
            self.follow_sort_desc = False
        self.sort_treeview(self.follow_tree, col, self.follow_sort_desc, self.refresh_follow_table)
    def open_advance_filter(self):
        win = tb.Toplevel(self.root)
        win.title("高级筛选")
        win.geometry("420x260")
        win.resizable(True,True)
        win.grab_set()
        tb.Label(win, text="客户阶段").pack(pady=(10,2))
        filter_stage = tb.Combobox(win, values=["全部"]+STAGE_LIST, state="readonly")
        filter_stage.set("全部")
        filter_stage.pack()
        tb.Label(win, text="标签模糊匹配").pack(pady=(10,2))
        filter_tag = tb.Entry(win)
        filter_tag.pack()
        def do_filter():
            s = filter_stage.get()
            t = filter_tag.get().strip()
            sql = '''
            SELECT DISTINCT c.id, c.name, c.legal_person, c.phone, c.address, c.stage, c.tags,
                (SELECT follow_date FROM follow_log WHERE cust_id = c.id ORDER BY id DESC LIMIT 1) AS last_date,
                (SELECT remark FROM follow_log WHERE cust_id = c.id ORDER BY id DESC LIMIT 1) AS last_remark
            FROM customer c WHERE 1=1
            '''
            params = []
            if s != "全部":
                sql += " AND c.stage = ?"
                params.append(s)
            if t:
                sql += " AND c.tags LIKE ?"
                params.append(f"%{t}%")
            sql += " ORDER BY c.id DESC"
            for i in self.cust_tree.get_children():
                self.cust_tree.delete(i)
            rows = db_query(sql, params)
            for row in rows:
                self.cust_tree.insert("", tk.END, values=row)
            win.destroy()
        tb.Button(win, text="执行筛选", command=do_filter, bootstyle=SUCCESS).pack(pady=15)
    def do_search(self):
        keyword = self.search_var.get().strip()
        if not keyword:
            self.refresh_cust_table()
            return
        sql_base = '''
            SELECT DISTINCT c.id, c.name, c.legal_person, c.phone, c.address, c.stage, c.tags,
                (SELECT follow_date FROM follow_log WHERE cust_id = c.id ORDER BY id DESC LIMIT 1) AS last_date,
                (SELECT remark FROM follow_log WHERE cust_id = c.id ORDER BY id DESC LIMIT 1) AS last_remark
            FROM customer c
            LEFT JOIN follow_log fl ON c.id = fl.cust_id
            WHERE
        '''
        cond_list = []
        for field in SEARCH_FIELDS:
            cond_list.append(f"c.{field} LIKE ?")
        cond_list.append("fl.remark LIKE ?")
        where_sql = " OR ".join(cond_list)
        final_sql = sql_base + where_sql + " ORDER BY c.id DESC"
        params = [f"%{keyword}%"] * len(cond_list)
        for i in self.cust_tree.get_children():
            self.cust_tree.delete(i)
        rows = db_query(final_sql, params)
        for row in rows:
            self.cust_tree.insert("", tk.END, values=row)
    def open_customer_detail(self, event):
        sel = self.cust_tree.selection()
        if not sel:
            return
        item = self.cust_tree.item(sel[0])
        cust_id = item["values"][0]
        cust_info = db_query("SELECT * FROM customer WHERE id = ?", (cust_id,))
        if not cust_info:
            return
        cust_info = cust_info[0]
        win, canvas, inner = create_scrolled_toplevel(self.root, f"客户详情 | ID:{cust_id}", w=940, h=660)
        btn_frame = tb.Frame(inner)
        btn_frame.pack(pady=6, padx=10, anchor="w")
        info_frame = tb.LabelFrame(inner, text="完整客户信息", padding=10)
        info_frame.pack(fill="x", padx=10, pady=6)
        db_fields = ["id","name","legal_person","phone","address","stage","business_scope","registered_cap","remark","tags"]
        field_map = {k:v for k,v in zip(db_fields, cust_info)}
        self.detail_vars = {}
        row_idx = 0
        visible_cfg_fields = [x for x in CFG["customer_detail_fields"] if x["visible"]]
        for cfg_item in visible_cfg_fields:
            f_name = cfg_item["field"]
            f_label = cfg_item["label"]
            tb.Label(info_frame, text=f_label).grid(row=row_idx, column=0, sticky="w", pady=3)
            val = field_map.get(f_name,"") if field_map.get(f_name,"") is not None else ""
            var = tk.StringVar(value=str(val))
            ent = tb.Entry(info_frame, textvariable=var, width=80)
            ent.grid(row=row_idx, column=1, padx=10, pady=3)
            self.detail_vars[f_name] = var
            row_idx += 1
        log_frame = tb.LabelFrame(inner, text="全部历史跟进记录", padding=10)
        log_frame.pack(expand=True, fill="both", padx=10, pady=6)
        log_tree = tb.Treeview(log_frame, columns=("ftype","fdate","next_dt","remark"), show="headings", height=8)
        log_tree.heading("ftype", text="跟进方式")
        log_tree.heading("fdate", text="跟进日期")
        log_tree.heading("next_dt", text="下次回访")
        log_tree.heading("remark", text="备注")
        log_tree.column("remark", width=460)
        log_tree.pack(fill="both", expand=True)
        log_rows = db_query("""
            SELECT follow_type, follow_date, next_remind, remark
            FROM follow_log WHERE cust_id=? ORDER BY follow_date DESC
        """, (cust_id,))
        for r in log_rows:
            log_tree.insert("", tk.END, values=r)
        def save_edit():
            update_data = []
            set_sql_parts = []
            for cfg_item in visible_cfg_fields:
                fn = cfg_item["field"]
                if fn == "id":
                    continue
                set_sql_parts.append(f"{fn}=?")
                update_data.append(self.detail_vars[fn].get())
            update_data.append(cust_id)
            sql = f"UPDATE customer SET {','.join(set_sql_parts)} WHERE id=?"
            db_execute(sql, tuple(update_data))
            Messagebox.show_info("信息修改成功！跟进记录不受影响", "完成")
            self.refresh_all()
            win.destroy()
        tb.Button(btn_frame, text="保存修改", command=save_edit, bootstyle=SUCCESS).pack(side="left", padx=6)
        tb.Button(btn_frame, text="关闭", command=win.destroy).pack(side="left", padx=6)
    def delete_customer(self):
        sel = self.cust_tree.selection()
        if not sel:
            Messagebox.show_warning("请先选中表格中的客户！")
            return
        item = self.cust_tree.item(sel[0])
        cust_id = item["values"][0]
        cust_name = item["values"][1]
        confirm = Messagebox.okcancel(
            f"确认删除【{cust_name}】？\n⚠️警告：会同步删除该客户全部跟进记录，无法恢复！",
            "危险操作"
        )
        if confirm not in ("OK", "确定"):
            return
        db_execute("DELETE FROM follow_log WHERE cust_id = ?", (cust_id,))
        db_execute("DELETE FROM customer WHERE id = ?", (cust_id,))
        self.refresh_all()
        Messagebox.show_info("删除完成！")
    def add_cust(self):
        name = self.var_cust_name.get().strip()
        if not name:
            Messagebox.show_warning("企业名称不能为空")
            return
        exist = db_query("SELECT id FROM customer WHERE name = ?", (name,))
        if len(exist) > 0:
            Messagebox.show_warning(f"客户【{name}】已存在")
            return
        data = (name, self.var_legal.get(), self.var_phone.get(), self.var_addr.get(), self.var_stage.get(), "", "", "", self.var_tag.get())
        db_execute('''INSERT INTO customer(name,legal_person,phone,address,stage,business_scope,registered_cap,remark,tags)
            VALUES (?,?,?,?,?,?,?,?,?)''', data)
        self.refresh_all()
        Messagebox.show_info("新增客户成功")
    def refresh_cust_table(self, sort_col=None, is_desc=False):
        for i in self.cust_tree.get_children():
            self.cust_tree.delete(i)
        order_by = "ORDER BY c.id DESC"
        if sort_col:
            col_map = {cfg["field"]: cfg["field"] for cfg in CFG["customer_table_columns"]}
            db_col = col_map.get(sort_col, "id")
            order = "DESC" if is_desc else "ASC"
            order_by = f"ORDER BY c.{db_col} {order}"
        sql = f"""
            SELECT c.id, c.name, c.legal_person, c.phone, c.address, c.stage, c.tags,
                   (SELECT follow_date FROM follow_log WHERE cust_id = c.id ORDER BY id DESC LIMIT 1) AS last_date,
                   (SELECT remark FROM follow_log WHERE cust_id = c.id ORDER BY id DESC LIMIT 1) AS last_remark
            FROM customer c {order_by}
        """
        rows = db_query(sql)
        for row in rows:
            self.cust_tree.insert("", tk.END, values=row)
    # ====================== XLSX导入向导：垂直+横向滚动 ======================
    def import_xlsx(self):
        win = tb.Toplevel(self.root)
        win.title("XLSX导入配置向导")
        win.geometry("1020x720")
        win.resizable(True, True)
        win.grab_set()
        win.transient(self.root)
        self.xlsx_path = None
        self.xlsx_sheets = []
        self.selected_sheet = tk.StringVar()
        self.header_row = tk.IntVar(value=1)
        self.xlsx_headers = []
        self.xlsx_preview_data = []
        self.field_mapping = {}
        # 默认不开启额外列导入，修改说明文本
        self.import_extra_cols = tk.BooleanVar(value=False)
        config_frame = tb.LabelFrame(win, text="1. 基础配置", padding=10)
        config_frame.pack(fill="x", padx=10, pady=5)
        tb.Button(config_frame, text="选择XLSX文件", command=self.select_xlsx_file, bootstyle=PRIMARY).grid(row=0, column=0, padx=5)
        self.file_label = tb.Label(config_frame, text="未选择文件", width=60)
        self.file_label.grid(row=0, column=1, padx=5)
        tb.Label(config_frame, text="选择Sheet:").grid(row=1, column=0, padx=5, pady=8)
        self.sheet_combo = tb.Combobox(config_frame, textvariable=self.selected_sheet, width=30, state="readonly")
        self.sheet_combo.grid(row=1, column=1, padx=5, pady=8)
        tb.Label(config_frame, text="表头行号(Excel行号):").grid(row=1, column=2, padx=5, pady=8)
        tb.Spinbox(config_frame, from_=1, to=100, textvariable=self.header_row, width=10).grid(row=1, column=3, padx=5, pady=8)
        tb.Button(config_frame, text="加载预览", command=self.load_xlsx_preview, bootstyle=SUCCESS).grid(row=1, column=4, padx=10, pady=8)
        preview_frame = tb.LabelFrame(win, text="2. 数据预览（前10行）", padding=10)
        preview_frame.pack(fill="both", expand=True, padx=10, pady=5)
        preview_container = tb.Frame(preview_frame)
        preview_container.pack(fill="both", expand=True)
        self.preview_tree = tb.Treeview(preview_container, show="headings", height=10)
        preview_v_scroll = tb.Scrollbar(preview_container, orient=VERTICAL, command=self.preview_tree.yview)
        preview_h_scroll = tb.Scrollbar(preview_container, orient=HORIZONTAL, command=self.preview_tree.xview)
        self.preview_tree.configure(yscrollcommand=preview_v_scroll.set, xscrollcommand=preview_h_scroll.set)
        preview_v_scroll.pack(side=RIGHT, fill=Y)
        preview_h_scroll.pack(side=BOTTOM, fill=X)
        self.preview_tree.pack(fill="both", expand=True)
        mapping_frame = tb.LabelFrame(win, text="3. 字段映射配置（已自动匹配，可手动修改）", padding=10)
        mapping_frame.pack(fill="x", padx=10, pady=5)
        canvas = tk.Canvas(mapping_frame, height=180)
        scrollbar = tb.Scrollbar(mapping_frame, orient=VERTICAL, command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=RIGHT, fill=Y)
        canvas.pack(side=LEFT, fill=BOTH, expand=True)
        self.mapping_inner = tb.Frame(canvas)
        canvas.create_window((0,0), window=self.mapping_inner, anchor="nw")
        def _on_map_conf(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
        self.mapping_inner.bind("<Configure>", _on_map_conf)
        def _mw(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _mw)
        tb.Label(mapping_frame, text="⚠️ 系统已自动匹配同名字段，您可以下拉调整映射关系，多列对应同一字段会用逗号合并").pack(pady=5)
        # 修改复选框说明，明确不会自动新增字段
        tb.Checkbutton(mapping_frame, text="导入Excel中额外的列（需手动配置字段后导入，不会自动新增数据库字段）", variable=self.import_extra_cols).pack(pady=5)
        btn_frame = tb.Frame(win)
        btn_frame.pack(pady=10)
        tb.Button(btn_frame, text="开始导入", command=self.do_xlsx_import, bootstyle=SUCCESS, width=15).pack(side="left", padx=10)
        tb.Button(btn_frame, text="关闭", command=win.destroy, bootstyle=SECONDARY, width=15).pack(side="left", padx=10)
        self.import_win = win
        self.mapping_canvas = canvas
    def select_xlsx_file(self):
        fp = filedialog.askopenfilename(filetypes=[("Excel文件", "*.xlsx *.xls")])
        if not fp:
            return
        self.xlsx_path = fp
        self.file_label.config(text=os.path.basename(fp))
        wb = load_workbook(fp, read_only=True, data_only=True)
        self.xlsx_sheets = wb.sheetnames
        self.sheet_combo["values"] = self.xlsx_sheets
        if self.xlsx_sheets:
            self.selected_sheet.set(self.xlsx_sheets[0])
        wb.close()
    def load_xlsx_preview(self):
        if not self.xlsx_path:
            Messagebox.show_warning("请先选择XLSX文件")
            return
        sheet_name = self.selected_sheet.get()
        if not sheet_name:
            Messagebox.show_warning("请选择Sheet")
            return
        header_row = self.header_row.get()
        if header_row < 1:
            Messagebox.show_warning("表头行号必须≥1")
            return
        wb = load_workbook(self.xlsx_path, read_only=True, data_only=True)
        ws = wb[sheet_name]
        header_cells = ws[header_row]
        self.xlsx_headers = [cell.value if cell.value is not None else f"列_{i+1}" for i, cell in enumerate(header_cells)]
        self.xlsx_preview_data = []
        for row in ws.iter_rows(min_row=header_row+1, max_row=header_row+10, values_only=True):
            self.xlsx_preview_data.append(row)
        wb.close()
        for i in self.preview_tree.get_children():
            self.preview_tree.delete(i)
        self.preview_tree["columns"] = [f"col_{i}" for i in range(len(self.xlsx_headers))]
        for i, header in enumerate(self.xlsx_headers):
            self.preview_tree.heading(f"col_{i}", text=header)
            self.preview_tree.column(f"col_{i}", width=120)
        for row in self.xlsx_preview_data:
            self.preview_tree.insert("", tk.END, values=row)
        field_alias_map = {
            "name": ["公司名称", "企业名称", "公司", "企业", "客户名称", "名称", "单位名称", "主体名称"],
            "legal_person": ["法人", "法定代表人", "法人姓名", "法人代表", "法人名称", "法定代表人姓名"],
            "phone": ["联系电话", "手机", "电话", "联系方式", "手机号码", "联系手机", "座机", "联系座机", "tel"],
            "address": ["地址", "经营地址", "注册地址", "公司地址", "办公地址", "所在地址", "详细地址", "经营场所"],
            "stage": ["客户阶段", "阶段", "状态", "客户状态", "跟进阶段", "所处阶段", "合作状态", "客户等级"],
            "tags": ["标签", "客户标签", "分类", "客户分类", "类型", "客户类型", "行业", "所属行业", "客户来源"],
            "business_scope": ["经营范围", "营业范围", "业务范围", "公司经营范围", "经营项目", "主营业务", "业务内容"],
            "registered_cap": ["注册资本", "注册资金", "认缴资本", "实缴资本", "出资额", "注册资本金", "认缴金额", "实缴金额"],
            "remark": ["内部备注", "备注", "说明", "备注信息", "客户备注", "跟进备注", "内部说明", "其他说明"]
        }
        for widget in self.mapping_inner.winfo_children():
            widget.destroy()
        system_fields = [x for x in CFG["customer_detail_fields"] if x["field"] != "id"]
        self.field_mapping.clear()
        for i, cfg_item in enumerate(system_fields):
            field_name = cfg_item["field"]
            label = cfg_item["label"]
            matched_col = "(不导入)"
            alias_list = field_alias_map.get(field_name, [label])
            for header in self.xlsx_headers:
                header_lower = str(header).lower()
                for alias in alias_list:
                    if alias.lower() in header_lower:
                        matched_col = header
                        break
                if matched_col != "(不导入)":
                    break
            tb.Label(self.mapping_inner, text=label).grid(row=i, column=0, sticky="w", pady=3)
            var = tk.StringVar(value=matched_col)
            combo = tb.Combobox(self.mapping_inner, textvariable=var, values=["(不导入)"] + self.xlsx_headers, width=30, state="readonly")
            combo.grid(row=i, column=1, padx=10, pady=3)
            self.field_mapping[field_name] = var
        self.mapping_canvas.update_idletasks()
        self.mapping_canvas.config(scrollregion=self.mapping_canvas.bbox("all"))
    def do_xlsx_import(self):
        if not self.xlsx_path:
            Messagebox.show_warning("请先选择XLSX文件")
            return
        sheet_name = self.selected_sheet.get()
        header_row = self.header_row.get()
        name_mapping = self.field_mapping.get("name")
        if not name_mapping or name_mapping.get() == "(不导入)":
            Messagebox.show_error("【企业名称/name】字段必须选择对应Excel列！")
            return
        wb = load_workbook(self.xlsx_path, read_only=True, data_only=True)
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(min_row=header_row+1, values_only=True))
        wb.close()
        db_cols = get_table_columns("customer")
        extra_cols = []
        # 额外列仅导入已存在的数据库字段，不会自动新增字段
        if self.import_extra_cols.get():
            for header in self.xlsx_headers:
                mapped = False
                for db_field, var in self.field_mapping.items():
                    if var.get() == header:
                        mapped = True
                        break
                if not mapped and header in db_cols:
                    extra_cols.append(header)
                elif not mapped and header not in db_cols:
                    # 提示用户额外的列需要手动配置
                    Messagebox.show_warning(f"Excel列【{header}】未在数据库中找到对应字段，已跳过导入，请先手动在config.json和数据库中配置该字段后再导入。")
        added = 0
        skipped = 0
        for row in all_rows:
            row_dict = dict(zip(self.xlsx_headers, row))
            name_val = row_dict.get(name_mapping.get(), "").strip()
            if not name_val:
                skipped += 1
                continue
            exist = db_query("SELECT id FROM customer WHERE name = ?", (name_val,))
            if len(exist) > 0:
                skipped += 1
                continue
            insert_data = {}
            for db_field, var in self.field_mapping.items():
                excel_col = var.get()
                if excel_col == "(不导入)":
                    insert_data[db_field] = ""
                    continue
                if db_field in insert_data:
                    insert_data[db_field] += f", {row_dict.get(excel_col, '')}"
                else:
                    insert_data[db_field] = str(row_dict.get(excel_col, "")).strip()
            if self.import_extra_cols.get():
                for col in extra_cols:
                    insert_data[col] = str(row_dict.get(col, "")).strip()
            insert_data.pop("id", None)
            fields_list = list(insert_data.keys())
            placeholders = ["?"] * len(fields_list)
            values_list = [insert_data[k] for k in fields_list]
            sql_insert = f"INSERT INTO customer ({','.join(fields_list)}) VALUES ({','.join(placeholders)})"
            db_execute(sql_insert, tuple(values_list))
            added += 1
        self.refresh_all()
        msg = f"导入完成：新增{added}条，跳过重复/空名称{skipped}条\n"
        if extra_cols:
            msg += f"⚠️ 已导入手动配置的额外字段：{','.join(extra_cols)}\n"
        Messagebox.show_info(msg)
        self.import_win.destroy()
    def export_all_csv(self):
        fp = filedialog.asksaveasfilename(
            defaultextension=".csv", filetypes=[("CSV", "*.csv")],
            initialfile=f"全部客户_{datetime.now().strftime('%Y%m%d')}.csv"
        )
        if not fp: return
        rows = db_query("""
            SELECT c.id, c.name, c.legal_person, c.phone, c.address, c.stage,c.tags,
                   (SELECT follow_date FROM follow_log WHERE cust_id=c.id ORDER BY id DESC LIMIT 1),
                   (SELECT remark FROM follow_log WHERE cust_id=c.id ORDER BY id DESC LIMIT 1)
            FROM customer c ORDER BY c.id
        """)
        headers = [x["label"] for x in CFG["customer_table_columns"] if x["visible"]]
        with open(fp, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(headers)
            w.writerows(rows)
        Messagebox.show_info(f"导出{len(rows)}条客户数据")
    def export_all_xlsx(self):
        fp = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")],
            initialfile=f"全部客户_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
        if not fp: return
        wb = Workbook()
        ws = wb.active
        ws.title="客户清单"
        headers = [x["label"] for x in CFG["customer_table_columns"] if x["visible"]]
        ws.append(headers)
        rows = db_query("""
            SELECT c.id, c.name, c.legal_person, c.phone, c.address, c.stage,c.tags,
                   (SELECT follow_date FROM follow_log WHERE cust_id=c.id ORDER BY id DESC LIMIT 1),
                   (SELECT remark FROM follow_log WHERE cust_id=c.id ORDER BY id DESC LIMIT 1)
            FROM customer c ORDER BY c.id
        """)
        for r in rows:
            ws.append(r)
        wb.save(fp)
        Messagebox.show_info("XLSX导出完成")
    def _export_by_date_common(self, title, filter_field, export_type):
        win = tb.Toplevel(self.root)
        win.title(title)
        win.geometry("360x220")
        win.resizable(True,True)
        win.grab_set()
        tb.Label(win, text="开始日期").pack(pady=(20, 2))
        date_start = tb.DateEntry(win, dateformat="%Y-%m-%d")
        date_start.pack(pady=2)
        tb.Label(win, text="结束日期").pack(pady=(10, 2))
        date_end = tb.DateEntry(win, dateformat="%Y-%m-%d")
        date_end.pack(pady=2)
        def do_export():
            d1 = date_start.entry.get()
            d2 = date_end.entry.get()
            if export_type == "csv":
                fp = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV", "*.csv")])
            else:
                fp = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
            if not fp: return
            rows = db_query(f"""
                SELECT f.id, c.name, f.follow_type, f.follow_date, f.next_remind, f.remark
                FROM follow_log f LEFT JOIN customer c ON f.cust_id=c.id
                WHERE f.{filter_field} BETWEEN ? AND ?
            """, (d1, d2))
            headers = ["日志ID","客户名称","跟进方式","跟进日期","下次回访","备注"]
            if export_type == "csv":
                with open(fp, "w", encoding="utf-8-sig", newline="") as f:
                    w = csv.writer(f)
                    w.writerow(headers)
                    w.writerows(rows)
            else:
                wb = Workbook()
                ws = wb.active
                ws.title="跟进记录"
                ws.append(headers)
                for r in rows:
                    ws.append(r)
                wb.save(fp)
            win.destroy()
            Messagebox.show_info(f"导出{len(rows)}条跟进记录")
        tb.Button(win, text="确认导出", command=do_export, bootstyle=SUCCESS).pack(pady=15)
    def export_by_date_csv(self):
        self._export_by_date_common("按跟进日期导出CSV", "follow_date", "csv")
    def export_by_date_xlsx(self):
        self._export_by_date_common("按跟进日期导出XLSX", "follow_date", "xlsx")
    def export_by_next_remind_csv(self):
        self._export_by_date_common("按下次回访导出CSV", "next_remind", "csv")
    def export_by_next_remind_xlsx(self):
        self._export_by_date_common("按下次回访导出XLSX", "next_remind", "xlsx")
    def build_follow_tab(self):
        f_in = tb.Frame(self.frame_follow)
        f_in.pack(fill="x", padx=8, pady=4)
        tb.Label(f_in, text="选择客户：").grid(row=0, column=0, padx=3)
        self.autocomplete_cust = AutocompleteEntry(f_in, width=26)
        self.autocomplete_cust.grid(row=0, column=1, padx=3)
        tb.Label(f_in, text="跟进方式").grid(row=0, column=2, padx=3)
        self.var_followtype = tk.StringVar()
        tb.Combobox(f_in, textvariable=self.var_followtype, values=FOLLOW_TYPE_LIST, width=14, state="readonly").grid(row=0, column=3, padx=3)
        tb.Label(f_in, text="跟进日期").grid(row=0, column=4, padx=3)
        self.date_follow = tb.DateEntry(f_in, dateformat="%Y-%m-%d", width=12)
        self.date_follow.grid(row=0, column=5, padx=3)
        tb.Label(f_in, text="下次回访").grid(row=0, column=6, padx=3)
        self.date_next = tb.DateEntry(f_in, dateformat="%Y-%m-%d", width=12)
        self.date_next.grid(row=0, column=7, padx=3)
        f_in2 = tb.Frame(self.frame_follow)
        f_in2.pack(fill="x", padx=8, pady=4)
        tb.Label(f_in2, text="备注").grid(row=0, column=0, padx=3)
        self.var_remark = tk.StringVar()
        tb.Entry(f_in2, textvariable=self.var_remark, width=66).grid(row=0, column=1, padx=3)
        tb.Button(f_in2, text="新增跟进记录", command=self.add_follow, bootstyle=SUCCESS).grid(row=0, column=2, padx=4)
        tb.Button(f_in2, text="编辑选中记录", command=self.edit_follow_log, bootstyle=INFO).grid(row=0, column=3, padx=(80,4))
        tb.Button(f_in2, text="删除选中记录", command=self.delete_follow_log, bootstyle=DANGER).grid(row=0, column=4, padx=4)
        log_cols_def = [x for x in CFG["follow_log_table_columns"] if x["visible"]]
        log_col_ids = [item["field"] for item in log_cols_def]
        
        # 跟进日志表格容器（增加滚动条）
        follow_table_container = tb.Frame(self.frame_follow)
        follow_table_container.pack(expand=True, fill="both", padx=8, pady=6)
        
        self.follow_tree = tb.Treeview(follow_table_container, columns=log_col_ids, show="headings")
        # 纵向滚动条
        follow_v_scroll = tb.Scrollbar(follow_table_container, orient=VERTICAL, command=self.follow_tree.yview)
        # 横向滚动条
        follow_h_scroll = tb.Scrollbar(follow_table_container, orient=HORIZONTAL, command=self.follow_tree.xview)
        self.follow_tree.configure(yscrollcommand=follow_v_scroll.set, xscrollcommand=follow_h_scroll.set)
        
        # 滚动条布局
        follow_v_scroll.pack(side=RIGHT, fill=Y)
        follow_h_scroll.pack(side=BOTTOM, fill=X)
        self.follow_tree.pack(fill="both", expand=True)
        
        # 配置列信息
        for cfg_col in log_cols_def:
            col_id = cfg_col["field"]
            self.follow_tree.heading(col_id, text=cfg_col["label"])
            self.follow_tree.column(col_id, width=cfg_col["width"])
            self.follow_tree.heading(col_id, command=lambda c=col_id: self.on_follow_sort(c))
        self.selected_cust_id = None
    def reload_customer_combobox(self):
        rows = db_query("SELECT id,name FROM customer ORDER BY name")
        data = []
        for cid,name in rows:
            data.append(f"{cid} | {name}")
        self.autocomplete_cust.set_items(data)
    def add_follow(self):
        sel_text = self.autocomplete_cust.get()
        if not sel_text or "|" not in sel_text:
            Messagebox.show_warning("请先在下拉框选择客户！")
            return
        cid = sel_text.split("|")[0].strip()
        data = (cid, self.var_followtype.get(), self.date_follow.entry.get(), self.date_next.entry.get(), self.var_remark.get())
        db_execute("INSERT INTO follow_log(cust_id,follow_type,follow_date,next_remind,remark) VALUES (?,?,?,?,?)", data)
        self.refresh_all()
        Messagebox.show_info("跟进记录保存成功")
    def edit_follow_log(self):
        sel = self.follow_tree.selection()
        if not sel:
            Messagebox.show_warning("请选中一条跟进记录")
            return
        item = self.follow_tree.item(sel[0])
        vals = item["values"]
        log_id = vals[0]
        old_cust_id = str(vals[1])
        old_cust_name = vals[2]
        old_follow_type = vals[3]
        old_follow_date = vals[4]
        old_next_remind = vals[5]
        old_remark = vals[6]
        win, canvas, inner = create_scrolled_toplevel(self.root, "编辑跟进记录", w=520, h=460)
        tb.Label(inner, text="所属客户(ID|名称)").pack(pady=(8,2))
        edit_cust_ac = AutocompleteEntry(inner, width=42)
        cust_list = db_query("SELECT id,name FROM customer ORDER BY name")
        item_list = [f"{cid} | {name}" for cid,name in cust_list]
        edit_cust_ac.set_items(item_list)
        edit_cust_ac.set(f"{old_cust_id} | {old_cust_name}")
        edit_cust_ac.pack()
        tb.Label(inner, text="跟进方式").pack(pady=(8,2))
        edit_type = tb.Combobox(inner, values=FOLLOW_TYPE_LIST, state="readonly")
        edit_type.set(old_follow_type)
        edit_type.pack()
        tb.Label(inner, text="跟进日期").pack(pady=(8,2))
        edit_fdate = tb.DateEntry(inner, dateformat="%Y-%m-%d")
        edit_fdate.entry.delete(0, tk.END)
        edit_fdate.entry.insert(0, old_follow_date)
        edit_fdate.pack()
        tb.Label(inner, text="下次回访").pack(pady=(8,2))
        edit_ndate = tb.DateEntry(inner, dateformat="%Y-%m-%d")
        edit_ndate.entry.delete(0, tk.END)
        edit_ndate.entry.insert(0, old_next_remind)
        edit_ndate.pack()
        tb.Label(inner, text="备注").pack(pady=(8,2))
        edit_remark = tb.Entry(inner, width=50)
        edit_remark.insert(0, old_remark)
        edit_remark.pack()
        def save():
            cust_text = edit_cust_ac.get()
            if not cust_text or "|" not in cust_text:
                Messagebox.show_warning("请选择客户！")
                return
            new_cid = cust_text.split("|")[0].strip()
            ft = edit_type.get()
            fd = edit_fdate.entry.get()
            nd = edit_ndate.entry.get()
            rm = edit_remark.get()
            db_execute("""
                UPDATE follow_log SET cust_id=?, follow_type=?,follow_date=?,next_remind=?,remark=? WHERE id=?
            """, (new_cid, ft, fd, nd, rm, log_id))
            self.refresh_all()
            win.destroy()
            Messagebox.show_info("修改完成")
        tb.Button(inner, text="保存", command=save, bootstyle=SUCCESS).pack(pady=12)
    def delete_follow_log(self):
        sel = self.follow_tree.selection()
        if not sel:
            Messagebox.show_warning("请选中一条跟进记录")
            return
        item = self.follow_tree.item(sel[0])
        log_id = item["values"][0]
        confirm = Messagebox.okcancel("确定删除这条跟进记录？不可恢复！")
        if confirm not in ("OK", "确定"):
            return
        db_execute("DELETE FROM follow_log WHERE id=?", (log_id,))
        self.refresh_all()
    def refresh_follow_table(self, sort_col=None, is_desc=False):
        for i in self.follow_tree.get_children():
            self.follow_tree.delete(i)
        order_by = "ORDER BY f.next_remind ASC"
        if sort_col:
            col_map = {cfg["field"]: cfg["field"] for cfg in CFG["follow_log_table_columns"]}
            db_col = col_map.get(sort_col, "id")
            order = "DESC" if is_desc else "ASC"
            order_by = f"ORDER BY f.{db_col} {order}"
        sql = f"""
            SELECT f.id AS fid, f.cust_id, c.name AS cust_name, f.follow_type, f.follow_date, f.next_remind, f.remark
            FROM follow_log f LEFT JOIN customer c ON f.cust_id = c.id {order_by}
        """
        rows = db_query(sql)
        for row in rows:
            self.follow_tree.insert("", tk.END, values=row)
    def check_remind(self):
        today = datetime.now().strftime("%Y-%m-%d")
        rows = db_query("SELECT DISTINCT c.name FROM follow_log f LEFT JOIN customer c ON f.cust_id=c.id WHERE f.next_remind=?", (today,))
        if rows:
            name_list = "\n".join([x[0] for x in rows])
            Messagebox.show_warning(f"【今日待回访客户】\n{name_list}", "回访提醒")
        self.root.after(1800000, self.check_remind)
if __name__ == "__main__":
    init_db()
    root = tb.Window(themename="cosmo")
    app = SalesCRM(root)
    root.mainloop()